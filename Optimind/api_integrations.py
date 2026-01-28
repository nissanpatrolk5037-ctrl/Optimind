import json
import re
import os
import asyncio
import subprocess
import zipfile
import shutil
import hashlib
import random
import string
import socket
import logging
from datetime import datetime, timedelta
from pathlib import Path
from email.message import EmailMessage
from email import message_from_bytes
import sys
from typing import Dict, Any, Optional, List, Union

# Import with error handling for optional packages
try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False
    print("Warning: requests module not installed. Some features will be unavailable.")

try:
    import discord
    DISCORD_AVAILABLE = True
except ImportError:
    DISCORD_AVAILABLE = False

try:
    import boto3
    BOTO3_AVAILABLE = True
except ImportError:
    BOTO3_AVAILABLE = False

try:
    import imaplib
    IMAPLIB_AVAILABLE = True
except ImportError:
    IMAPLIB_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import sqlite3
    SQLITE_AVAILABLE = True
except ImportError:
    SQLITE_AVAILABLE = False

try:
    import mysql.connector
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False

try:
    import psycopg2
    PSYCOPG2_AVAILABLE = True
except ImportError:
    PSYCOPG2_AVAILABLE = False

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

try:
    import paramiko
    PARAMIKO_AVAILABLE = True
except ImportError:
    PARAMIKO_AVAILABLE = False

try:
    import pyautogui
    PYAUTOGUI_AVAILABLE = True
except ImportError:
    PYAUTOGUI_AVAILABLE = False

try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

try:
    import pyperclip
    PYPERCLIP_AVAILABLE = True
except ImportError:
    PYPERCLIP_AVAILABLE = False

try:
    import speech_recognition as sr
    SPEECHRECOGNITION_AVAILABLE = True
except ImportError:
    SPEECHRECOGNITION_AVAILABLE = False

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import qrcode
    QRCODE_AVAILABLE = True
except ImportError:
    QRCODE_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

try:
    from cryptography.fernet import Fernet
    CRYPTOGRAPHY_AVAILABLE = True
except ImportError:
    CRYPTOGRAPHY_AVAILABLE = False

try:
    from django.core.mail import send_mail as dj_send_mail
    DJANGO_AVAILABLE = True
except ImportError:
    DJANGO_AVAILABLE = False

# Additional imports for new integrations
try:
    import tweepy
    TWITTER_AVAILABLE = True
except ImportError:
    TWITTER_AVAILABLE = False

try:
    import dropbox
    DROPBOX_AVAILABLE = True
except ImportError:
    DROPBOX_AVAILABLE = False

try:
    import docker
    DOCKER_AVAILABLE = True
except ImportError:
    DOCKER_AVAILABLE = False

try:
    import wikipedia
    WIKIPEDIA_AVAILABLE = True
except ImportError:
    WIKIPEDIA_AVAILABLE = False

try:
    import pywhatkit as kit
    PYWHATKIT_AVAILABLE = True
except ImportError:
    PYWHATKIT_AVAILABLE = False

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import nmap
    NMAP_AVAILABLE = True
except ImportError:
    NMAP_AVAILABLE = False

try:
    import yfinance
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False

try:
    # Try importing from Optimind if available
    from auto_coder import get_llm_response as groq_answer
    from data import *
    OPTIMIND_AVAILABLE = True
except ImportError:
    OPTIMIND_AVAILABLE = False
    # Define fallback functions
    def groq_answer(prompt: str) -> str:
        return "Optimind not available. Please install Optimind package."
    
    # Define fallback environment variables
    DISCORD_TOKEN = os.environ.get('DISCORD_TOKEN', '')
    WHATSAPP_TOKEN = os.environ.get('WHATSAPP_TOKEN', '')
    WHATSAPP_PHONE_ID = os.environ.get('WHATSAPP_PHONE_ID', '')
    GOOGLE_TRANSLATE_API_KEY = os.environ.get('GOOGLE_TRANSLATE_API_KEY', '')
    SLACK_TOKEN = os.environ.get('SLACK_TOKEN', '')
    TWITTER_API_KEY = os.environ.get('TWITTER_API_KEY', '')
    TWITTER_API_SECRET = os.environ.get('TWITTER_API_SECRET', '')
    TWITTER_ACCESS_TOKEN = os.environ.get('TWITTER_ACCESS_TOKEN', '')
    TWITTER_ACCESS_SECRET = os.environ.get('TWITTER_ACCESS_SECRET', '')
    GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN', '')
    DROPBOX_TOKEN = os.environ.get('DROPBOX_TOKEN', '')
    OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '')
    OPENWEATHER_API_KEY = os.environ.get('OPENWEATHER_API_KEY', '')
    BITLY_TOKEN = os.environ.get('BITLY_TOKEN', '')

try:
    from google.oauth2.credentials import Credentials
    from googleapiclient.http import MediaFileUpload
    from googleapiclient.discovery import build
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False

# Create a session for requests if available
if REQUESTS_AVAILABLE:
    SESSION = requests.Session()
else:
    SESSION = None

IS_WINDOWS = sys.platform.startswith("win")
IS_MACOS = sys.platform == "darwin"
IS_LINUX = sys.platform.startswith("linux")

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def extract_params(query: str, service_type: str) -> Dict[str, str]:
    """Generic parameter extraction using NLP"""
    prompts = {
        "aws_control": "Extract 'action', 'instance_id', 'region' from query. Return as JSON.",
        "upload_to_google_drive": "Extract 'path', 'filename' from query. Return as JSON.",
        "schedule_google_calendar_event": "Extract 'title', 'start_time', 'end_time' from query. Return as JSON.",
        "send_discord_message": "Extract 'message', 'channel_id' from query. Return as JSON.",
        "send_whatsapp_message": "Extract 'phone', 'message' from query. Return as JSON.",
        "send_email": "Extract 'message', 'recipient_email' from query. Return as JSON.",
        "send_multiple_emails": "Extract 'message', 'emails', 'sender' from query. Return as JSON.",
        "download_email_attachments": "Extract 'email_user', 'email_pass', 'folder' from query. Return as JSON.",
        "send_slack_message": "Extract 'message', 'channel', 'username' from query. Return as JSON.",
        "post_tweet": "Extract 'tweet_text', 'image_path' from query. Return as JSON.",
        "github_create_repo": "Extract 'repo_name', 'description', 'is_private' from query. Return as JSON.",
        "azure_vm_control": "Extract 'action', 'vm_name', 'resource_group' from query. Return as JSON.",
        "dropbox_upload": "Extract 'local_path', 'remote_path' from query. Return as JSON.",
        "docker_container_control": "Extract 'action', 'container_name', 'image_name' from query. Return as JSON.",
        "openai_chat": "Extract 'prompt', 'model', 'max_tokens' from query. Return as JSON.",
        "translate_text": "Extract 'text', 'target_language', 'source_language' from query. Return as JSON.",
        "wikipedia_search": "Extract 'query', 'sentences' from query. Return as JSON.",
        "youtube_download": "Extract 'url', 'format', 'output_path' from query. Return as JSON.",
        "qr_code_generate": "Extract 'data', 'filename', 'size' from query. Return as JSON.",
        "pdf_create": "Extract 'content', 'filename', 'title' from query. Return as JSON.",
        "screenshot": "Extract 'filename', 'region' from query. Return as JSON.",
        "image_ocr": "Extract 'image_path', 'api_key' from query. Return as JSON.",
        "video_convert": "Extract 'input_path', 'output_path', 'format' from query. Return as JSON.",
        "audio_extract": "Extract 'video_path', 'output_path' from query. Return as JSON.",
        "excel_operations": "Extract 'operation', 'file_path', 'sheet_name' from query. Return as JSON.",
        "database_query": "Extract 'query', 'database_type', 'host' from query. Return as JSON.",
        "data_visualization": "Extract 'data_file', 'chart_type', 'output_file' from query. Return as JSON.",
        "weather_info": "Extract 'city', 'country', 'units' from query. Return as JSON.",
        "stock_price": "Extract 'symbol', 'period', 'interval' from query. Return as JSON.",
        "currency_convert": "Extract 'amount', 'from_currency', 'to_currency' from query. Return as JSON.",
        "encrypt_decrypt_file": "Extract 'operation', 'input_file', 'output_file', 'key' from query. Return as JSON.",
        "password_generate": "Extract 'length', 'include_symbols', 'include_numbers' from query. Return as JSON.",
        "url_shorten": "Extract 'long_url', 'custom_alias' from query. Return as JSON.",
        "domain_lookup": "Extract 'domain_name', 'info_type' from query. Return as JSON.",
        "hash_generate": "Extract 'text', 'algorithm' from query. Return as JSON.",
        "system_info": "Extract 'info_type' from query. Return as JSON.",
        "process_control": "Extract 'action', 'process_name', 'pid' from query. Return as JSON.",
        "backup_files": "Extract 'source', 'destination', 'compression' from query. Return as JSON.",
        "network_scan": "Extract 'target', 'port_range' from query. Return as JSON.",
        "ssh_execute": "Extract 'host', 'command', 'username', 'password' from query. Return as JSON.",
        "web_scrape": "Extract 'url', 'element', 'output' from query. Return as JSON.",
        "file_operations": "Extract 'operation', 'path', 'destination' from query. Return as JSON.",
        "git_operations": "Extract 'operation', 'repo_path', 'branch' from query. Return as JSON.",
        "api_test": "Extract 'url', 'method', 'payload' from query. Return as JSON.",
        "log_analyze": "Extract 'log_file', 'pattern', 'output' from query. Return as JSON.",
    }
    
    if service_type in prompts:
        try:
            prompt = f"{prompts[service_type]} Query: '{query}'. If a parameter is not mentioned, set it to 'none'. Return ONLY JSON, no explanations."
            nlp_result = groq_answer(prompt)
            try:
                return json.loads(nlp_result)
            except json.JSONDecodeError:
                # Fallback regex extraction
                pattern = r'"([^"]+)":\s*"([^"]*)"'
                matches = re.findall(pattern, nlp_result)
                return {k: v for k, v in matches} if matches else {}
        except Exception as e:
            logger.error(f"Error in extract_params: {e}")
            return {}
    return {}

def get_language_name(lang_code: str) -> str:
    """Convert language code to full name"""
    language_names = {
        'af': 'Afrikaans', 'sq': 'Albanian', 'am': 'Amharic', 'ar': 'Arabic',
        'hy': 'Armenian', 'az': 'Azerbaijani', 'eu': 'Basque', 'be': 'Belarusian',
        'bn': 'Bengali', 'bs': 'Bosnian', 'bg': 'Bulgarian', 'ca': 'Catalan',
        'ceb': 'Cebuano', 'ny': 'Chichewa', 'zh-cn': 'Chinese (Simplified)',
        'zh-tw': 'Chinese (Traditional)', 'co': 'Corsican', 'hr': 'Croatian',
        'cs': 'Czech', 'da': 'Danish', 'nl': 'Dutch', 'en': 'English',
        'eo': 'Esperanto', 'et': 'Estonian', 'tl': 'Filipino', 'fi': 'Finnish',
        'fr': 'French', 'fy': 'Frisian', 'gl': 'Galician', 'ka': 'Georgian',
        'de': 'German', 'el': 'Greek', 'gu': 'Gujarati', 'ht': 'Haitian Creole',
        'ha': 'Hausa', 'haw': 'Hawaiian', 'he': 'Hebrew', 'hi': 'Hindi',
        'hmn': 'Hmong', 'hu': 'Hungarian', 'is': 'Icelandic', 'ig': 'Igbo',
        'id': 'Indonesian', 'ga': 'Irish', 'it': 'Italian', 'ja': 'Japanese',
        'jw': 'Javanese', 'kn': 'Kannada', 'kk': 'Kazakh', 'km': 'Khmer',
        'ko': 'Korean', 'ku': 'Kurdish (Kurmanji)', 'ky': 'Kyrgyz', 'lo': 'Lao',
        'la': 'Latin', 'lv': 'Latvian', 'lt': 'Lithuanian', 'lb': 'Luxembourgish',
        'mk': 'Macedonian', 'mg': 'Malagasy', 'ms': 'Malay', 'ml': 'Malayalam',
        'mt': 'Maltese', 'mi': 'Maori', 'mr': 'Marathi', 'mn': 'Mongolian',
        'my': 'Myanmar (Burmese)', 'ne': 'Nepali', 'no': 'Norwegian',
        'ps': 'Pashto', 'fa': 'Persian', 'pl': 'Polish', 'pt': 'Portuguese',
        'pa': 'Punjabi', 'ro': 'Romanian', 'ru': 'Russian', 'sm': 'Samoan',
        'gd': 'Scots Gaelic', 'sr': 'Serbian', 'st': 'Sesotho', 'sn': 'Shona',
        'sd': 'Sindhi', 'si': 'Sinhala', 'sk': 'Slovak', 'sl': 'Slovenian',
        'so': 'Somali', 'es': 'Spanish', 'su': 'Sundanese', 'sw': 'Swahili',
        'sv': 'Swedish', 'tg': 'Tajik', 'ta': 'Tamil', 'te': 'Telugu',
        'th': 'Thai', 'tr': 'Turkish', 'uk': 'Ukrainian', 'ur': 'Urdu',
        'uz': 'Uzbek', 'vi': 'Vietnamese', 'cy': 'Welsh', 'xh': 'Xhosa',
        'yi': 'Yiddish', 'yo': 'Yoruba', 'zu': 'Zulu'
    }
    return language_names.get(lang_code.lower(), lang_code)

def _get_status_description(status_code: int) -> str:
    """Get description for HTTP status codes"""
    status_descriptions = {
        100: "Continue",
        101: "Switching Protocols",
        200: "OK",
        201: "Created",
        202: "Accepted",
        204: "No Content",
        301: "Moved Permanently",
        302: "Found",
        304: "Not Modified",
        400: "Bad Request",
        401: "Unauthorized",
        403: "Forbidden",
        404: "Not Found",
        405: "Method Not Allowed",
        409: "Conflict",
        422: "Unprocessable Entity",
        429: "Too Many Requests",
        500: "Internal Server Error",
        502: "Bad Gateway",
        503: "Service Unavailable",
        504: "Gateway Timeout"
    }
    return status_descriptions.get(status_code, "Unknown Status")

def _get_log_level_emoji(level: str) -> str:
    """Get emoji for log level"""
    emoji_map = {
        'ERROR': '❌',
        'FATAL': '💀',
        'CRITICAL': '🔥',
        'WARN': '⚠️',
        'WARNING': '⚠️',
        'INFO': 'ℹ️',
        'DEBUG': '🐛'
    }
    return emoji_map.get(level, '📝')

def _get_status_emoji(status_code: str) -> str:
    """Get emoji for HTTP status code"""
    try:
        code = int(status_code)
        if 200 <= code < 300:
            return '✅'
        elif 300 <= code < 400:
            return '🔄'
        elif 400 <= code < 500:
            return '⚠️'
        elif 500 <= code < 600:
            return '❌'
        else:
            return '❓'
    except:
        return '❓'

# ========== ORIGINAL FUNCTIONS WITH ERROR HANDLING ==========

def aws_control(query: Optional[str] = None) -> str:
    """Control AWS EC2 instances"""
    try:
        if not BOTO3_AVAILABLE:
            return "❌ Boto3 not installed. Run: pip install boto3"
        
        nlp = extract_params(query, "aws_control") if query else {}
        
        action = nlp.get("action", "none") if nlp.get("action") != "none" else input("Type in the Action (start/stop/status): ")
        instance_id = nlp.get("instance_id", "none") if nlp.get("instance_id") != "none" else input("Type in the Instance ID: ")
        region = nlp.get("region", "none") if nlp.get("region") != "none" else "us-east-1"
        
        ec2 = boto3.client("ec2", region_name=region)
        
        if action.lower() == "start":
            ec2.start_instances(InstanceIds=[instance_id])
            return f"✅ EC2 instance {instance_id} started in {region}"
        elif action.lower() == "stop":
            ec2.stop_instances(InstanceIds=[instance_id])
            return f"✅ EC2 instance {instance_id} stopped in {region}"
        elif action.lower() == "status":
            response = ec2.describe_instances(InstanceIds=[instance_id])
            state = response['Reservations'][0]['Instances'][0]['State']['Name']
            return f"📊 EC2 instance {instance_id} status: {state}"
        else:
            return "❌ Action must be 'start', 'stop', or 'status'"
            
    except Exception as e:
        logger.error(f"AWS control error: {e}")
        return f"❌ Failed to control EC2 instance: {str(e)}"

def upload_to_google_drive(query: Optional[str] = None) -> str:
    """Upload file to Google Drive"""
    try:
        if not GOOGLE_API_AVAILABLE:
            return "❌ Google API client not installed. Run: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib"
        
        nlp = extract_params(query, "upload_to_google_drive") if query else {}
        
        path = nlp.get("path", "none") if nlp.get("path") != "none" else input("Enter the file path: ")
        custom_name = nlp.get("filename", "none") if nlp.get("filename") != "none" else input("Enter custom filename (press Enter for original): ") or None
        
        if not os.path.exists(path):
            return "❌ File not found"
        
        filename = custom_name or os.path.basename(path)
        
        if not os.path.exists("google_token.json"):
            return "❌ google_token.json not found. Please authenticate with Google first."
        
        creds = Credentials.from_authorized_user_file(
            "google_token.json", ["https://www.googleapis.com/auth/drive"]
        )
        service = build("drive", "v3", credentials=creds)
        
        media = MediaFileUpload(path, resumable=True)
        file = service.files().create(
            body={"name": filename},
            media_body=media
        ).execute()
        
        return f"✅ Uploaded '{filename}' to Google Drive. File ID: {file.get('id', 'No ID returned')}"
        
    except Exception as e:
        logger.error(f"Google Drive upload error: {e}")
        return f"❌ Failed to upload file: {str(e)}"

def send_discord_message(query: Optional[str] = None) -> str:
    """Send Discord message"""
    try:
        if not DISCORD_AVAILABLE:
            return "❌ Discord.py not installed. Run: pip install discord.py"
        
        nlp = extract_params(query, "send_discord_message") if query else {}
        
        message = nlp.get("message", "none") if nlp.get("message") != "none" else input("Enter the message to send: ")
        channel_id = int(nlp.get("channel_id", "none")) if nlp.get("channel_id") != "none" else int(input("Enter the channel ID to send to: "))
        
        if not message:
            return "❌ Message is required."
        
        if not DISCORD_TOKEN:
            return "❌ DISCORD_TOKEN not set in environment variables."
        
        async def _discord_send():
            intents = discord.Intents.default()
            client = discord.Client(intents=intents)
            
            @client.event
            async def on_ready():
                try:
                    channel = client.get_channel(channel_id)
                    if channel:
                        await channel.send(message)
                        print(f"✅ Discord message sent to channel {channel_id}")
                    else:
                        print(f"❌ Channel {channel_id} not found")
                except Exception as e:
                    print(f"Discord send failed: {e}")
                finally:
                    await client.close()
            
            try:
                await client.start(DISCORD_TOKEN)
            except Exception as e:
                print(f"Discord client failed to start: {e}")
        
        asyncio.run(_discord_send())
        return f"✅ Discord message sent to channel {channel_id}"
        
    except Exception as e:
        logger.error(f"Discord message error: {e}")
        return f"❌ Failed to send Discord message: {str(e)}"

def image_ocr(query: Optional[str] = None) -> str:
    """Perform OCR on an image using OCR.Space API"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed. Run: pip install requests"
        
        nlp = extract_params(query, "image_ocr") if query else {}
        
        image_path = nlp.get("image_path", "none") if nlp.get("image_path") != "none" else input("Enter image path: ")
        api_key = nlp.get("api_key", "none") if nlp.get("api_key") != "none" else "K85328613788957"
        
        if not image_path:
            return "❌ Error: No image path provided."
        
        p = Path(image_path)
        if not p.exists():
            return "❌ Error: Image not found."
        
        with p.open("rb") as img_file:
            r = SESSION.post(
                "https://api.ocr.space/parse/image",
                files={"image": img_file},
                data={"apikey": api_key, "language": "eng", "OCREngine": "2"},
                timeout=60
            )
        r.raise_for_status()
        
        try:
            result = r.json()
        except ValueError:
            return f"❌ Response not JSON: {r.text[:200]}..."
        
        if result.get("IsErroredOnProcessing"):
            return "❌ OCR Failed: " + str(result.get("ErrorMessage"))
        
        parsed = result.get("ParsedResults")
        if parsed and parsed[0].get("ParsedText"):
            return parsed[0]["ParsedText"].strip()
        
        return "⚠️ No text found in image."
        
    except Exception as e:
        logger.error(f"OCR error: {e}")
        return f"❌ OCR request failed: {str(e)}"

def git_operations(query: Optional[str] = None) -> str:
    """Perform Git operations (clone, pull, push, commit)"""
    try:
        nlp = extract_params(query, "git_operations") if query else {}
        
        operation = nlp.get("operation", "none") if nlp.get("operation") != "none" else input("Git operation (clone/pull/push/commit/status/branch/log): ")
        repo_path = nlp.get("repo_path", "none") if nlp.get("repo_path") != "none" else input("Repository path: ")
        branch = nlp.get("branch", "none") if nlp.get("branch") != "none" else "main"
        commit_message = nlp.get("commit_message", "none") if nlp.get("commit_message") != "none" else None
        
        # Try to import git
        try:
            import git
        except ImportError:
            return "❌ GitPython not installed. Run: pip install gitpython"
        
        if operation == "clone":
            url = input("Repository URL: ")
            print(f"Cloning {url} to {repo_path}...")
            git.Repo.clone_from(url, repo_path)
            return f"✅ Repository cloned to {repo_path}"
        
        elif operation == "status":
            if not os.path.exists(repo_path):
                return "❌ Repository path does not exist"
            
            repo = git.Repo(repo_path)
            status = repo.git.status()
            changes = repo.git.diff('HEAD')
            
            result = f"📊 Git Status for {repo_path}:\n"
            result += f"Branch: {repo.active_branch}\n"
            result += f"Status:\n{status}\n"
            
            if changes:
                result += f"Changes:\n{changes[:500]}..."  # Limit output
            else:
                result += "No changes to commit."
            
            return result
        
        # ... (rest of git operations with proper error handling)
        
        else:
            return "❌ Invalid operation. Use: clone, status, pull, push, commit, branch, log"
            
    except git.InvalidGitRepositoryError:
        return f"❌ Not a valid Git repository: {repo_path}"
    except git.GitCommandError as e:
        return f"❌ Git command failed: {e}"
    except Exception as e:
        logger.error(f"Git operations error: {e}")
        return f"❌ Git operation failed: {str(e)}"

def translate_text(query: Optional[str] = None) -> str:
    """Translate text between languages"""
    try:
        nlp = extract_params(query, "translate_text") if query else {}
        
        text = nlp.get("text", "none") if nlp.get("text") != "none" else input("Text to translate: ")
        target_lang = nlp.get("target_language", "none") if nlp.get("target_language") != "none" else input("Target language (e.g., es, fr, de, ja): ")
        source_lang = nlp.get("source_language", "none") if nlp.get("source_language") != "none" else "auto"
        
        # Try googletrans first
        try:
            from googletrans import Translator
            translator = Translator()
            result = translator.translate(text, dest=target_lang, src=source_lang)
            
            result_str = f"🌐 Translation:\n"
            result_str += f"📝 Original ({result.src}): {text}\n"
            result_str += f"🔤 Translated ({target_lang}): {result.text}\n"
            
            if hasattr(result, 'pronunciation') and result.pronunciation:
                result_str += f"🗣️ Pronunciation: {result.pronunciation}\n"
            
            return result_str
            
        except ImportError:
            # Fallback to Google Translate API
            if not GOOGLE_TRANSLATE_API_KEY:
                return "❌ Google Translate API key not set. Install googletrans: pip install googletrans==4.0.0rc1"
            
            if not REQUESTS_AVAILABLE:
                return "❌ Requests module required. Run: pip install requests"
            
            import json
            from urllib.parse import quote
            
            encoded_text = quote(text)
            url = f"https://translation.googleapis.com/language/translate/v2"
            params = {
                'q': text,
                'target': target_lang,
                'format': 'text',
                'key': GOOGLE_TRANSLATE_API_KEY
            }
            
            if source_lang != 'auto':
                params['source'] = source_lang
            
            response = requests.post(url, params=params)
            
            if response.status_code == 200:
                data = response.json()
                translation = data['data']['translations'][0]['translatedText']
                detected_lang = data['data']['translations'][0].get('detectedSourceLanguage', 'auto')
                
                result_str = f"🌐 Translation:\n"
                result_str += f"📝 Original ({detected_lang}): {text}\n"
                result_str += f"🔤 Translated ({target_lang}): {translation}\n"
                return result_str
            else:
                return f"❌ Translation failed: {response.text}"
                
    except Exception as e:
        logger.error(f"Translation error: {e}")
        return f"❌ Translation failed: {str(e)}"

def pdf_create(query: Optional[str] = None) -> str:
    """Create PDF files from text, images, or HTML"""
    try:
        if not FPDF_AVAILABLE:
            return "❌ FPDF not installed. Run: pip install fpdf"
        
        nlp = extract_params(query, "pdf_create") if query else {}
        
        content = nlp.get("content", "none") if nlp.get("content") != "none" else input("Enter content (text, HTML, or image path): ")
        filename = nlp.get("filename", "none") if nlp.get("filename") != "none" else f"document_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        title = nlp.get("title", "none") if nlp.get("title") != "none" else "Generated Document"
        content_type = nlp.get("content_type", "none") if nlp.get("content_type") != "none" else "auto"
        
        # Ensure filename has .pdf extension
        if not filename.lower().endswith('.pdf'):
            filename += '.pdf'
        
        # Determine content type
        if content_type == "auto":
            if content.lower().startswith('<html') or content.lower().startswith('<!doctype'):
                content_type = "html"
            elif os.path.exists(content) and content.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                content_type = "image"
            else:
                content_type = "text"
        
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        
        if content_type == "text":
            # Add title
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, txt=title, ln=1, align='C')
            pdf.set_font("Arial", size=12)
            pdf.ln(10)
            
            # Add content
            lines = content.split('\n')
            for line in lines:
                if line.strip():
                    pdf.multi_cell(0, 10, line)
            
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(200, 10, txt=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=1, align='C')
        
        elif content_type == "image":
            if not PIL_AVAILABLE:
                return "❌ PIL/Pillow required for image PDFs. Run: pip install pillow"
            
            if not os.path.exists(content):
                return f"❌ Image file not found: {content}"
            
            # Add image
            pdf.image(content, x=10, y=30, w=190)
            pdf.ln(100)
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(200, 10, txt=f"Image: {os.path.basename(content)}", ln=1, align='C')
        
        elif content_type == "html":
            try:
                # Try with xhtml2pdf
                from xhtml2pdf import pisa
                import tempfile
                
                html_content = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <title>{title}</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 40px; }}
                        h1 {{ color: #333; }}
                        p {{ line-height: 1.6; }}
                    </style>
                </head>
                <body>
                    <h1>{title}</h1>
                    {content}
                    <p style="font-style: italic; text-align: center; margin-top: 50px;">
                        Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}
                    </p>
                </body>
                </html>
                """
                
                with open(filename, 'wb') as pdf_file:
                    pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)
                
                if pisa_status.err:
                    return f"❌ HTML to PDF conversion failed"
                
            except ImportError:
                return "❌ xhtml2pdf not installed. Run: pip install xhtml2pdf"
        
        else:
            return f"❌ Unsupported content type: {content_type}"
        
        if content_type != "html":  # HTML PDFs are saved in the try block
            pdf.output(filename)
        
        file_size = os.path.getsize(filename) / 1024  # KB
        
        result = f"✅ PDF created successfully!\n"
        result += f"📄 File: {filename}\n"
        result += f"📏 Size: {file_size:.1f} KB\n"
        result += f"📋 Type: {content_type}\n"
        
        return result
        
    except Exception as e:
        logger.error(f"PDF creation error: {e}")
        return f"❌ PDF creation failed: {str(e)}"

def docker_container_control(query: Optional[str] = None) -> str:
    """Control Docker containers"""
    try:
        if not DOCKER_AVAILABLE:
            return "❌ Docker Python SDK not installed. Run: pip install docker"
        
        nlp = extract_params(query, "docker_container_control") if query else {}
        
        action = nlp.get("action", "none") if nlp.get("action") != "none" else input("Action (start/stop/restart/run/list/logs): ")
        container_name = nlp.get("container_name", "none") if nlp.get("container_name") != "none" else None
        image_name = nlp.get("image_name", "none") if nlp.get("image_name") != "none" else None
        
        client = docker.from_env()
        
        if action == "list":
            containers = client.containers.list(all=True)
            
            if not containers:
                return "📦 No Docker containers found"
            
            result = f"🐳 Docker Containers ({len(containers)}):\n"
            for container in containers:
                status = "🟢" if container.status == "running" else "🔴"
                result += f"{status} {container.name[:20]:20} {container.short_id:12} {container.status}\n"
            return result
        
        elif action == "run" and image_name:
            print(f"🚀 Running {image_name}...")
            container = client.containers.run(image_name, detach=True)
            return f"✅ Container started: {container.name} ({container.short_id})"
        
        elif action in ["start", "stop", "restart"] and container_name:
            container = client.containers.get(container_name)
            getattr(container, action)()
            return f"✅ Container {container_name} {action}ed"
        
        elif action == "logs" and container_name:
            container = client.containers.get(container_name)
            logs = container.logs(tail=50).decode('utf-8')
            return f"📄 Logs for {container_name}:\n{logs[:2000]}"
        
        else:
            return "❌ Invalid action or missing parameters"
            
    except docker.errors.NotFound:
        return "❌ Container not found"
    except docker.errors.APIError as e:
        return f"❌ Docker API error: {e}"
    except Exception as e:
        logger.error(f"Docker control error: {e}")
        return f"❌ Docker operation failed: {str(e)}"

def excel_operations(query: Optional[str] = None) -> str:
    """Excel file operations"""
    try:
        if not PANDAS_AVAILABLE:
            return "❌ Pandas not installed. Run: pip install pandas openpyxl"
        
        nlp = extract_params(query, "excel_operations") if query else {}
        
        operation = nlp.get("operation", "none") if nlp.get("operation") != "none" else input("Operation (read/write/create): ")
        file_path = nlp.get("file_path", "none") if nlp.get("file_path") != "none" else input("Excel file path: ")
        sheet_name = nlp.get("sheet_name", "none") if nlp.get("sheet_name") != "none" else "Sheet1"
        
        if operation == "read":
            if not os.path.exists(file_path):
                return f"❌ File not found: {file_path}"
            
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return f"📊 Excel Data ({len(df)} rows):\n{df.head().to_string()}"
        
        elif operation == "write":
            # Create sample data
            data = {"Name": ["Alice", "Bob", "Charlie"], "Score": [85, 92, 78]}
            df = pd.DataFrame(data)
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
            return f"✅ Excel file created: {file_path}"
        
        elif operation == "create":
            df = pd.DataFrame()
            df.to_excel(file_path, index=False)
            return f"✅ Empty Excel file created: {file_path}"
        
        else:
            return "❌ Invalid operation"
            
    except Exception as e:
        logger.error(f"Excel operations error: {e}")
        return f"❌ Excel operation failed: {str(e)}"

def api_test(self, query: Optional[str] = None) -> str:
    """Test API endpoints with history tracking"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed. Run: pip install requests"
        
        nlp = self.extract_params(query, "api_test") if query else {}
        
        url = nlp.get("url", "none") if nlp.get("url") != "none" else input("Enter API URL: ")
        method = nlp.get("method", "none") if nlp.get("method") != "none" else "GET"
        payload = nlp.get("payload", "none") if nlp.get("payload") != "none" else None
        
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        import time
        start_time = time.time()
        
        # Make the API request
        if method.upper() == 'GET':
            response = requests.get(url, timeout=10)
        elif method.upper() == 'POST':
            response = requests.post(url, json=payload if payload else None, timeout=10)
        elif method.upper() == 'PUT':
            response = requests.put(url, json=payload if payload else None, timeout=10)
        elif method.upper() == 'DELETE':
            response = requests.delete(url, timeout=10)
        else:
            return f"❌ Unsupported method: {method}"
        
        response_time = (time.time() - start_time) * 1000
        
        # Create test record
        test_record = {
            'timestamp': datetime.now().isoformat(),
            'url': url,
            'method': method.upper(),
            'status': response.status_code,
            'time': response_time,
            'success': 200 <= response.status_code < 300
        }
        
        # ===== IN-MEMORY HISTORY =====
        if not hasattr(self, 'api_test_history'):
            self.api_test_history = []
        
        self.api_test_history.append(test_record)
        
        # Keep only last 50 entries in memory
        if len(self.api_test_history) > 50:
            self.api_test_history = self.api_test_history[-50:]
        
        # ===== FILE-BASED HISTORY (AUTO-CREATES FILE) =====
        history_file = 'api_test_history.json'
        file_history = []
        
        try:
            # Try to load existing history
            if os.path.exists(history_file):
                with open(history_file, 'r') as f:
                    file_history = json.load(f)
            
            # Add new record
            file_history.append(test_record)
            
            # Keep only last 50 entries
            if len(file_history) > 50:
                file_history = file_history[-50:]
            
            # Save to file (auto-creates if doesn't exist)
            with open(history_file, 'w') as f:
                json.dump(file_history, f, indent=2)
                
        except Exception as e:
            # If file operations fail, just continue without saving to file
            logger.warning(f"Could not save API test history to file: {e}")
        
        # ===== BUILD RESULT STRING =====
        result = f"🔧 API Test Results\n"
        result += "=" * 60 + "\n"
        result += f"🌐 URL: {url}\n"
        result += f"📡 Method: {method.upper()}\n"
        result += f"⏱️  Response Time: {response_time:.2f} ms\n"
        result += f"📊 Status: {response.status_code} {response.reason}\n"
        
        # Status emoji
        if 200 <= response.status_code < 300:
            result += "✅ Success\n"
        elif 400 <= response.status_code < 500:
            result += "⚠️  Client Error\n"
        else:
            result += "❌ Server Error\n"
        
        # Show response body
        try:
            json_data = response.json()
            result += f"\n📄 Response Body (JSON):\n{json.dumps(json_data, indent=2)[:500]}"
            if len(json.dumps(json_data)) > 500:
                result += "..."
        except:
            if response.text:
                result += f"\n📄 Response Body (Text):\n{response.text[:500]}"
                if len(response.text) > 500:
                    result += "..."
            else:
                result += "\n📄 Response Body: [Empty]"
        
        return result
        
    except requests.exceptions.Timeout:
        return "❌ Request timeout after 10 seconds"
    
    except requests.exceptions.ConnectionError:
        return "❌ Connection failed. Check URL and network connection."
    
    except Exception as e:
        logger.error(f"API test error: {e}")
        return f"❌ API test failed: {str(e)[:200]}"

def log_analyze(query: Optional[str] = None) -> str:
    """Analyze log files"""
    try:
        nlp = extract_params(query, "log_analyze") if query else {}
        
        log_file = nlp.get("log_file", "none") if nlp.get("log_file") != "none" else input("Enter log file path: ")
        pattern = nlp.get("pattern", "none") if nlp.get("pattern") != "none" else None
        
        if not os.path.exists(log_file):
            return f"❌ Log file not found: {log_file}"
        
        with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        total_lines = len(lines)
        error_count = 0
        warning_count = 0
        info_count = 0
        
        for line in lines:
            line_lower = line.lower()
            if 'error' in line_lower:
                error_count += 1
            elif 'warning' in line_lower:
                warning_count += 1
            elif 'info' in line_lower:
                info_count += 1
        
        result = f"📊 Log Analysis: {os.path.basename(log_file)}\n"
        result += "=" * 60 + "\n"
        result += f"📈 Total Lines: {total_lines:,}\n"
        result += f"❌ Errors: {error_count:,}\n"
        result += f"⚠️  Warnings: {warning_count:,}\n"
        result += f"ℹ️  Info: {info_count:,}\n"
        
        if pattern:
            matches = [line for line in lines if pattern in line]
            result += f"\n🔍 Pattern '{pattern}' found {len(matches)} times\n"
            if matches:
                result += "Sample matches:\n"
                for match in matches[:3]:
                    result += f"  • {match.strip()[:100]}\n"
        
        # Show last 5 lines
        result += f"\n📄 Last 5 lines:\n"
        for line in lines[-5:]:
            result += f"  {line.strip()}\n"
        
        return result
        
    except Exception as e:
        logger.error(f"Log analysis error: {e}")
        return f"❌ Log analysis failed: {str(e)}"

def file_operations(query: Optional[str] = None) -> str:
    """File operations (copy, move, delete, rename)"""
    try:
        nlp = extract_params(query, "file_operations") if query else {}
        
        operation = nlp.get("operation", "none") if nlp.get("operation") != "none" else input("Operation (copy/move/delete/rename/list): ")
        path = nlp.get("path", "none") if nlp.get("path") != "none" else input("Path: ")
        
        if operation == "list":
            if not os.path.exists(path):
                return f"❌ Path not found: {path}"
            
            if os.path.isdir(path):
                files = os.listdir(path)
                result = f"📁 Contents of {path}:\n"
                for file in files[:20]:  # Limit to 20 items
                    full_path = os.path.join(path, file)
                    if os.path.isdir(full_path):
                        result += f"📁 {file}/\n"
                    else:
                        size = os.path.getsize(full_path)
                        result += f"📄 {file} ({size:,} bytes)\n"
                if len(files) > 20:
                    result += f"... and {len(files) - 20} more items\n"
                return result
            else:
                return "❌ Path is not a directory"
        
        elif operation == "delete":
            if not os.path.exists(path):
                return f"❌ Path not found: {path}"
            
            if os.path.isdir(path):
                shutil.rmtree(path)
            else:
                os.remove(path)
            return f"✅ Deleted: {path}"
        
        elif operation in ["copy", "move", "rename"]:
            destination = nlp.get("destination", "none") if nlp.get("destination") != "none" else input("Destination: ")
            
            if not os.path.exists(path):
                return f"❌ Source not found: {path}"
            
            if operation == "copy":
                if os.path.isdir(path):
                    shutil.copytree(path, destination)
                else:
                    shutil.copy2(path, destination)
                return f"✅ Copied to: {destination}"
            
            elif operation == "move":
                shutil.move(path, destination)
                return f"✅ Moved to: {destination}"
            
            elif operation == "rename":
                os.rename(path, destination)
                return f"✅ Renamed to: {destination}"
        
        else:
            return "❌ Invalid operation"
            
    except Exception as e:
        logger.error(f"File operations error: {e}")
        return f"❌ File operation failed: {str(e)}"

def system_info(query: Optional[str] = None) -> str:
    """Get system information"""
    try:
        if not PSUTIL_AVAILABLE:
            return "❌ psutil not installed. Run: pip install psutil"
        
        import platform
        
        nlp = extract_params(query, "system_info") if query else {}
        info_type = nlp.get("info_type", "none") if nlp.get("info_type") != "none" else "all"
        
        result = "💻 System Information\n"
        result += "=" * 60 + "\n"
        
        if info_type in ["os", "all"]:
            result += f"🏷️ OS: {platform.system()} {platform.release()}\n"
            result += f"🐍 Python: {platform.python_version()}\n"
        
        if info_type in ["cpu", "all"]:
            cpu_percent = psutil.cpu_percent(interval=1)
            cpu_count = psutil.cpu_count()
            result += f"⚡ CPU: {cpu_count} cores, {cpu_percent}% usage\n"
        
        if info_type in ["memory", "all"]:
            memory = psutil.virtual_memory()
            result += f"💾 Memory: {memory.percent}% used ({memory.used/1e9:.1f} GB / {memory.total/1e9:.1f} GB)\n"
        
        if info_type in ["disk", "all"]:
            disk = psutil.disk_usage('/')
            result += f"💽 Disk: {disk.percent}% used ({disk.used/1e9:.1f} GB / {disk.total/1e9:.1f} GB)\n"
        
        if info_type in ["network", "all"]:
            net_io = psutil.net_io_counters()
            result += f"🌐 Network: Sent {net_io.bytes_sent/1e6:.1f} MB, Received {net_io.bytes_recv/1e6:.1f} MB\n"
        
        return result
        
    except Exception as e:
        logger.error(f"System info error: {e}")
        return f"❌ System info failed: {str(e)}"

def backup_files(query: Optional[str] = None) -> str:
    """Backup files/folders"""
    try:
        nlp = extract_params(query, "backup_files") if query else {}
        
        source = nlp.get("source", "none") if nlp.get("source") != "none" else input("Source path: ")
        destination = nlp.get("destination", "none") if nlp.get("destination") != "none" else input("Destination path: ")
        compression = nlp.get("compression", "none") if nlp.get("compression") != "none" else "zip"
        
        if not os.path.exists(source):
            return f"❌ Source not found: {source}"
        
        if os.path.isfile(source):
            if compression == "zip":
                with zipfile.ZipFile(destination, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(source, os.path.basename(source))
                return f"✅ File backed up to {destination}"
            else:
                shutil.copy2(source, destination)
                return f"✅ File copied to {destination}"
        
        elif os.path.isdir(source):
            if compression == "zip":
                with zipfile.ZipFile(destination, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for root, dirs, files in os.walk(source):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, source)
                            zipf.write(file_path, arcname)
                return f"✅ Folder backed up to {destination}"
            else:
                shutil.copytree(source, destination)
                return f"✅ Folder copied to {destination}"
        
        else:
            return "❌ Source not found"
            
    except Exception as e:
        logger.error(f"Backup error: {e}")
        return f"❌ Backup failed: {str(e)}"

def process_control(query: Optional[str] = None) -> str:
    """Control system processes"""
    try:
        if not PSUTIL_AVAILABLE:
            return "❌ psutil not installed. Run: pip install psutil"
        
        nlp = extract_params(query, "process_control") if query else {}
        
        action = nlp.get("action", "none") if nlp.get("action") != "none" else input("Action (list/kill): ")
        
        if action == "list":
            processes = []
            for proc in psutil.process_iter(['pid', 'name', 'cpu_percent']):
                try:
                    processes.append(f"PID: {proc.info['pid']}, Name: {proc.info['name']}, CPU: {proc.info['cpu_percent']}%")
                except:
                    pass
            
            result = f"📋 Running Processes (showing top 10 by CPU):\n"
            # Sort by CPU and show top 10
            sorted_procs = sorted(processes, key=lambda x: float(x.split('CPU: ')[1].replace('%', '')), reverse=True)[:10]
            result += "\n".join(sorted_procs)
            return result
        
        elif action == "kill":
            pid = nlp.get("pid", "none") if nlp.get("pid") != "none" else int(input("Process PID: "))
            try:
                psutil.Process(pid).kill()
                return f"✅ Process {pid} killed"
            except psutil.NoSuchProcess:
                return f"❌ Process {pid} not found"
        
        else:
            return "❌ Invalid action"
            
    except Exception as e:
        logger.error(f"Process control error: {e}")
        return f"❌ Process control failed: {str(e)}"

def screenshot(query: Optional[str] = None) -> str:
    """Take screenshot"""
    try:
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI not installed. Run: pip install pyautogui"
        
        nlp = extract_params(query, "screenshot") if query else {}
        
        filename = nlp.get("filename", "none") if nlp.get("filename") != "none" else f"screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        screenshot = pyautogui.screenshot()
        screenshot.save(filename)
        
        return f"✅ Screenshot saved as {filename}"
        
    except Exception as e:
        logger.error(f"Screenshot error: {e}")
        return f"❌ Screenshot failed: {str(e)}"

def password_generate(query: Optional[str] = None) -> str:
    """Generate secure password"""
    try:
        nlp = extract_params(query, "password_generate") if query else {}
        
        length = int(nlp.get("length", "none")) if nlp.get("length") != "none" else 12
        include_symbols = nlp.get("include_symbols", "none") != "no" if nlp.get("include_symbols") != "none" else True
        include_numbers = nlp.get("include_numbers", "none") != "no" if nlp.get("include_numbers") != "none" else True
        
        chars = string.ascii_letters
        if include_numbers:
            chars += string.digits
        if include_symbols:
            chars += "!@#$%^&*()"
        
        password = ''.join(random.choice(chars) for _ in range(length))
        
        # Check password strength
        strength = "Weak"
        if length >= 12 and include_numbers and include_symbols:
            strength = "Strong"
        elif length >= 8:
            strength = "Medium"
        
        return f"🔐 Generated Password ({strength}): {password}"
        
    except Exception as e:
        logger.error(f"Password generation error: {e}")
        return f"❌ Password generation failed: {str(e)}"

def hash_generate(query: Optional[str] = None) -> str:
    """Generate hash of text"""
    try:
        nlp = extract_params(query, "hash_generate") if query else {}
        
        text = nlp.get("text", "none") if nlp.get("text") != "none" else input("Text to hash: ")
        algorithm = nlp.get("algorithm", "none") if nlp.get("algorithm") != "none" else "sha256"
        
        if algorithm == "md5":
            hash_obj = hashlib.md5(text.encode())
        elif algorithm == "sha1":
            hash_obj = hashlib.sha1(text.encode())
        elif algorithm == "sha256":
            hash_obj = hashlib.sha256(text.encode())
        elif algorithm == "sha512":
            hash_obj = hashlib.sha512(text.encode())
        else:
            return "❌ Unsupported algorithm. Use: md5, sha1, sha256, sha512"
        
        hash_result = hash_obj.hexdigest()
        return f"🔒 {algorithm.upper()} Hash:\n{hash_result}"
        
    except Exception as e:
        logger.error(f"Hash generation error: {e}")
        return f"❌ Hash generation failed: {str(e)}"



def encrypt_decrypt_file(query: Optional[str] = None) -> str:
    """Encrypt or decrypt a file"""
    try:
        if not CRYPTOGRAPHY_AVAILABLE:
            return "❌ Cryptography not installed. Run: pip install cryptography"
        
        nlp = extract_params(query, "encrypt_decrypt_file") if query else {}
        
        operation = nlp.get("operation", "none") if nlp.get("operation") != "none" else input("Operation (encrypt/decrypt): ")
        input_file = nlp.get("input_file", "none") if nlp.get("input_file") != "none" else input("Input file: ")
        output_file = nlp.get("output_file", "none") if nlp.get("output_file") != "none" else input("Output file: ")
        key = nlp.get("key", "none") if nlp.get("key") != "none" else Fernet.generate_key()
        
        if not os.path.exists(input_file):
            return f"❌ Input file not found: {input_file}"
        
        cipher = Fernet(key)
        
        with open(input_file, 'rb') as f:
            data = f.read()
        
        if operation == "encrypt":
            encrypted = cipher.encrypt(data)
            with open(output_file, 'wb') as f:
                f.write(encrypted)
            return f"✅ File encrypted: {output_file}\n🔑 Key: {key.decode()}"
        
        elif operation == "decrypt":
            try:
                decrypted = cipher.decrypt(data)
                with open(output_file, 'wb') as f:
                    f.write(decrypted)
                return f"✅ File decrypted: {output_file}"
            except:
                return "❌ Decryption failed. Invalid key or file."
        
        else:
            return "❌ Invalid operation. Use 'encrypt' or 'decrypt'"
            
    except Exception as e:
        logger.error(f"Encryption/decryption error: {e}")
        return f"❌ Encryption/decryption failed: {str(e)}"

# ========== ADDITIONAL FUNCTIONS WITH ERROR HANDLING ==========

def detect_language(query: Optional[str] = None) -> str:
    """Detect language of text"""
    try:
        nlp = extract_params(query, "detect_language") if query else {}
        
        text = nlp.get("text", "none") if nlp.get("text") != "none" else input("Text to analyze: ")
        
        try:
            from googletrans import Translator
            translator = Translator()
            result = translator.detect(text)
            
            result_str = f"🔍 Language Detection:\n"
            result_str += f"📝 Text: {text}\n"
            result_str += f"🌐 Language: {result.lang} ({get_language_name(result.lang)})\n"
            result_str += f"📊 Confidence: {result.confidence*100:.1f}%\n"
            return result_str
            
        except ImportError:
            return "❌ Install: pip install googletrans==4.0.0rc1"
            
    except Exception as e:
        logger.error(f"Language detection error: {e}")
        return f"❌ Language detection failed: {str(e)}"

def youtube_download(query: Optional[str] = None) -> str:
    """Download YouTube video"""
    try:
        if not PYWHATKIT_AVAILABLE:
            return "❌ PyWhatKit not installed. Run: pip install pywhatkit"
        
        nlp = extract_params(query, "youtube_download") if query else {}
        
        url = nlp.get("url", "none") if nlp.get("url") != "none" else input("YouTube URL: ")
        format_type = nlp.get("format", "none") if nlp.get("format") != "none" else "mp4"
        output_path = nlp.get("output_path", "none") if nlp.get("output_path") != "none" else "downloads/"
        
        os.makedirs(output_path, exist_ok=True)
        
        if format_type == "mp3":
            kit.download_audio(url, output_path=output_path)
            return f"✅ Audio downloaded to {output_path}"
        else:
            kit.download_video(url, output_path=output_path)
            return f"✅ Video downloaded to {output_path}"
            
    except Exception as e:
        logger.error(f"YouTube download error: {e}")
        return f"❌ YouTube download failed: {str(e)}"

def audio_extract(query: Optional[str] = None) -> str:
    """Extract audio from video"""
    try:
        nlp = extract_params(query, "audio_extract") if query else {}
        
        video_path = nlp.get("video_path", "none") if nlp.get("video_path") != "none" else input("Video path: ")
        output_path = nlp.get("output_path", "none") if nlp.get("output_path") != "none" else f"{os.path.splitext(video_path)[0]}.mp3"
        
        if not os.path.exists(video_path):
            return "❌ Video file not found"
        
        # Try moviepy first
        try:
            from moviepy import VideoFileClip
            video = VideoFileClip(video_path)
            audio = video.audio
            audio.write_audiofile(output_path)
            video.close()
            return f"✅ Audio extracted: {output_path}"
            
        except ImportError:
            # Fallback to ffmpeg
            import subprocess
            cmd = [
                'ffmpeg', '-i', video_path,
                '-vn', '-acodec', 'mp3',
                '-ab', '192k', '-ar', '44100',
                '-y', output_path
            ]
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                return f"✅ Audio extracted: {output_path}"
            else:
                return f"❌ Audio extraction failed. Install ffmpeg or moviepy"
                
    except Exception as e:
        logger.error(f"Audio extraction error: {e}")
        return f"❌ Audio extraction failed: {str(e)}"

def video_convert(query: Optional[str] = None) -> str:
    """Convert video format"""
    try:
        nlp = extract_params(query, "video_convert") if query else {}
        
        input_path = nlp.get("input_path", "none") if nlp.get("input_path") != "none" else input("Input video path: ")
        output_path = nlp.get("output_path", "none") if nlp.get("output_path") != "none" else f"{os.path.splitext(input_path)[0]}_converted.mp4"
        
        if not os.path.exists(input_path):
            return "❌ Input video not found"
        
        # Use ffmpeg
        import subprocess
        cmd = [
            'ffmpeg', '-i', input_path,
            '-c:v', 'libx264', '-preset', 'medium',
            '-crf', '23', '-c:a', 'aac',
            '-b:a', '128k', '-y', output_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            return f"✅ Video converted: {output_path}"
        else:
            return f"❌ Video conversion failed. Install ffmpeg"
            
    except Exception as e:
        logger.error(f"Video conversion error: {e}")
        return f"❌ Video conversion failed: {str(e)}"

def weather_info(query: Optional[str] = None) -> str:
    """Get weather information"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        nlp = extract_params(query, "weather_info") if query else {}
        
        city = nlp.get("city", "none") if nlp.get("city") != "none" else input("City: ")
        country = nlp.get("country", "none") if nlp.get("country") != "none" else ""
        units = nlp.get("units", "none") if nlp.get("units") != "none" else "metric"
        
        if not OPENWEATHER_API_KEY:
            return "❌ OPENWEATHER_API_KEY not set in environment variables"
        
        location = f"{city},{country}" if country else city
        url = f"http://api.openweathermap.org/data/2.5/weather?q={location}&appid={OPENWEATHER_API_KEY}&units={units}"
        
        response = requests.get(url)
        
        if response.status_code == 200:
            data = response.json()
            temp = data['main']['temp']
            desc = data['weather'][0]['description']
            humidity = data['main']['humidity']
            wind = data['wind']['speed']
            
            return f"🌤️ Weather in {city}:\n🌡️ Temperature: {temp}°C\n☁️ Condition: {desc}\n💧 Humidity: {humidity}%\n💨 Wind: {wind} m/s"
        else:
            return f"❌ Weather data not found: {response.json().get('message', 'Unknown error')}"
            
    except Exception as e:
        logger.error(f"Weather info error: {e}")
        return f"❌ Weather API failed: {str(e)}"

def stock_price(query: Optional[str] = None) -> str:
    """Get stock price information"""
    try:
        if not YFINANCE_AVAILABLE:
            return "❌ yfinance not installed. Run: pip install yfinance"
        
        nlp = extract_params(query, "stock_price") if query else {}
        
        symbol = nlp.get("symbol", "none") if nlp.get("symbol") != "none" else input("Stock symbol (e.g., AAPL): ")
        period = nlp.get("period", "none") if nlp.get("period") != "none" else "1d"
        
        import yfinance as yf
        stock = yf.Ticker(symbol)
        hist = stock.history(period=period)
        
        if not hist.empty:
            current = hist['Close'].iloc[-1]
            prev_close = hist['Close'].iloc[0]
            change = current - prev_close
            pct_change = (change / prev_close) * 100
            
            return f"📈 {symbol} Stock:\n💰 Current: ${current:.2f}\n📊 Change: ${change:.2f} ({pct_change:.2f}%)"
        else:
            return f"❌ No data for {symbol}"
            
    except Exception as e:
        logger.error(f"Stock price error: {e}")
        return f"❌ Stock data failed: {str(e)}"

def currency_convert(query: Optional[str] = None) -> str:
    """Convert currency"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        nlp = extract_params(query, "currency_convert") if query else {}
        
        amount = float(nlp.get("amount", "none")) if nlp.get("amount") != "none" else float(input("Amount: "))
        from_curr = nlp.get("from_currency", "none") if nlp.get("from_currency") != "none" else input("From currency (e.g., USD): ")
        to_curr = nlp.get("to_currency", "none") if nlp.get("to_currency") != "none" else input("To currency (e.g., EUR): ")
        
        # Using free currency API
        url = f"https://api.exchangerate-api.com/v4/latest/{from_curr}"
        response = requests.get(url)
        
        if response.status_code == 200:
            data = response.json()
            rate = data['rates'].get(to_curr)
            
            if rate:
                converted = amount * rate
                return f"💱 {amount} {from_curr} = {converted:.2f} {to_curr}\n📈 Rate: 1 {from_curr} = {rate:.4f} {to_curr}"
            else:
                return f"❌ Currency {to_curr} not found"
        else:
            return f"❌ Currency API failed"
            
    except Exception as e:
        logger.error(f"Currency conversion error: {e}")
        return f"❌ Currency conversion failed: {str(e)}"

def url_shorten(query: Optional[str] = None) -> str:
    """Shorten URL"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        nlp = extract_params(query, "url_shorten") if query else {}
        
        long_url = nlp.get("long_url", "none") if nlp.get("long_url") != "none" else input("URL to shorten: ")
        
        # Using TinyURL API
        url = "https://tinyurl.com/api-create.php"
        params = {"url": long_url}
        
        response = requests.get(url, params=params)
        
        if response.status_code == 200:
            return f"🔗 Shortened URL: {response.text}"
        else:
            return f"❌ URL shortening failed"
            
    except Exception as e:
        logger.error(f"URL shorten error: {e}")
        return f"❌ URL shortening failed: {str(e)}"

def domain_lookup(query: Optional[str] = None) -> str:
    """Look up domain information"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        nlp = extract_params(query, "domain_lookup") if query else {}
        
        domain_name = nlp.get("domain_name", "none") if nlp.get("domain_name") != "none" else input("Domain name: ")
        
        # Clean domain
        domain_name = domain_name.strip()
        if domain_name.startswith(('http://', 'https://')):
            domain_name = domain_name.split('//')[1]
        if domain_name.startswith('www.'):
            domain_name = domain_name[4:]
        
        result = f"🔍 Domain Information for: {domain_name}\n"
        
        # Try WHOIS if available
        try:
            import whois
            domain_info = whois.whois(domain_name)
            
            if domain_info.domain_name:
                result += f"🌐 Domain: {domain_info.domain_name}\n"
            if domain_info.registrar:
                result += f"🏢 Registrar: {domain_info.registrar}\n"
            if domain_info.creation_date:
                result += f"📅 Created: {domain_info.creation_date}\n"
                
        except ImportError:
            result += "ℹ️ Install python-whois for more details\n"
        
        # Check if domain is reachable
        try:
            socket.gethostbyname(domain_name)
            result += "✅ Domain is reachable\n"
        except socket.gaierror:
            result += "❌ Domain not found\n"
        
        return result
        
    except Exception as e:
        logger.error(f"Domain lookup error: {e}")
        return f"❌ Domain lookup failed: {str(e)}"

def web_scrape(query: Optional[str] = None) -> str:
    """Scrape website content"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        nlp = extract_params(query, "web_scrape") if query else {}
        
        url = nlp.get("url", "none") if nlp.get("url") != "none" else input("URL: ")
        
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        # Simple text extraction
        from bs4 import BeautifulSoup
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Remove scripts and styles
        for script in soup(["script", "style"]):
            script.decompose()
        
        text = soup.get_text()
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = '\n'.join(chunk for chunk in chunks if chunk)
        
        return f"🌐 Scraped Content (first 500 chars):\n{text[:500]}..."
        
    except ImportError:
        return "❌ BeautifulSoup not installed. Run: pip install beautifulsoup4"
    except Exception as e:
        logger.error(f"Web scrape error: {e}")
        return f"❌ Web scraping failed: {str(e)}"

def data_visualization(query: Optional[str] = None) -> str:
    """Create data visualization"""
    try:
        if not MATPLOTLIB_AVAILABLE or not PANDAS_AVAILABLE:
            return "❌ Required: pip install matplotlib pandas"
        
        nlp = extract_params(query, "data_visualization") if query else {}
        
        data_file = nlp.get("data_file", "none") if nlp.get("data_file") != "none" else input("Data file (CSV): ")
        chart_type = nlp.get("chart_type", "none") if nlp.get("chart_type") != "none" else "line"
        output_file = nlp.get("output_file", "none") if nlp.get("output_file") != "none" else f"chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        if not os.path.exists(data_file):
            return f"❌ Data file not found: {data_file}"
        
        df = pd.read_csv(data_file)
        
        plt.figure(figsize=(10, 6))
        
        if chart_type == "line":
            for column in df.columns[1:]:  # Skip first column if it's index/date
                plt.plot(df.iloc[:, 0], df[column], label=column)
            plt.legend()
        elif chart_type == "bar":
            plt.bar(df.iloc[:, 0], df.iloc[:, 1])
        elif chart_type == "scatter":
            plt.scatter(df.iloc[:, 0], df.iloc[:, 1])
        elif chart_type == "histogram":
            plt.hist(df.iloc[:, 0], bins=20)
        
        plt.title(f"{chart_type.title()} Chart")
        plt.xlabel(df.columns[0])
        if len(df.columns) > 1:
            plt.ylabel(df.columns[1])
        plt.grid(True)
        plt.savefig(output_file)
        plt.close()
        
        return f"✅ Chart saved as {output_file}"
        
    except Exception as e:
        logger.error(f"Data visualization error: {e}")
        return f"❌ Visualization failed: {str(e)}"

def database_query(query: Optional[str] = None) -> str:
    """Execute database query"""
    try:
        nlp = extract_params(query, "database_query") if query else {}
        
        db_query = nlp.get("query", "none") if nlp.get("query") != "none" else input("SQL query: ")
        db_type = nlp.get("database_type", "none") if nlp.get("database_type") != "none" else "sqlite"
        
        if db_type == "sqlite":
            if not SQLITE_AVAILABLE:
                return "❌ SQLite not available"
            
            db_file = input("SQLite database file: ")
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            cursor.execute(db_query)
            
            if db_query.strip().upper().startswith("SELECT"):
                results = cursor.fetchall()
                columns = [description[0] for description in cursor.description]
                result_str = f"📋 Query Results ({len(results)} rows):\n"
                result_str += f"Columns: {', '.join(columns)}\n"
                for row in results[:10]:  # Show first 10 rows
                    result_str += f"{row}\n"
                if len(results) > 10:
                    result_str += f"... and {len(results) - 10} more rows\n"
                return result_str
            else:
                conn.commit()
                return f"✅ Query executed successfully. Rows affected: {cursor.rowcount}"
            
            conn.close()
            
        else:
            return f"❌ Database type {db_type} not implemented"
            
    except Exception as e:
        logger.error(f"Database query error: {e}")
        return f"❌ Database query failed: {str(e)}"

def network_scan(query: Optional[str] = None) -> str:
    """Scan network for devices"""
    try:
        if not NMAP_AVAILABLE:
            return "❌ python-nmap not installed. Run: pip install python-nmap"
        
        nlp = extract_params(query, "network_scan") if query else {}
        
        target = nlp.get("target", "none") if nlp.get("target") != "none" else "192.168.1.1"
        
        nm = nmap.PortScanner()
        nm.scan(hosts=target, arguments='-sn')  # Simple ping scan
        
        results = []
        for host in nm.all_hosts():
            if nm[host].state() == "up":
                results.append(f"🖥️ {host} is up")
        
        if results:
            return f"🌐 Network Scan Results:\n" + "\n".join(results)
        else:
            return "❌ No hosts found or scan failed"
            
    except Exception as e:
        logger.error(f"Network scan error: {e}")
        return f"❌ Network scan failed: {str(e)}"

def ssh_execute(query: Optional[str] = None) -> str:
    """Execute command via SSH"""
    try:
        if not PARAMIKO_AVAILABLE:
            return "❌ Paramiko not installed. Run: pip install paramiko"
        
        nlp = extract_params(query, "ssh_execute") if query else {}
        
        host = nlp.get("host", "none") if nlp.get("host") != "none" else input("Host: ")
        command = nlp.get("command", "none") if nlp.get("command") != "none" else input("Command: ")
        username = nlp.get("username", "none") if nlp.get("username") != "none" else input("Username: ")
        password = nlp.get("password", "none") if nlp.get("password") != "none" else input("Password: ")
        
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(host, username=username, password=password, timeout=10)
        
        stdin, stdout, stderr = client.exec_command(command)
        output = stdout.read().decode()
        error = stderr.read().decode()
        
        client.close()
        
        if output:
            return f"🖥️ SSH Output:\n{output[:1000]}"
        elif error:
            return f"❌ SSH Error:\n{error}"
        else:
            return "✅ Command executed (no output)"
            
    except Exception as e:
        logger.error(f"SSH execute error: {e}")
        return f"❌ SSH failed: {str(e)}"

def openai_chat(query: Optional[str] = None) -> str:
    """Chat with OpenAI GPT"""
    try:
        if not OPENAI_AVAILABLE:
            return "❌ OpenAI not installed. Run: pip install openai"
        
        nlp = extract_params(query, "openai_chat") if query else {}
        
        prompt = nlp.get("prompt", "none") if nlp.get("prompt") != "none" else input("Enter your prompt: ")
        model = nlp.get("model", "none") if nlp.get("model") != "none" else "gpt-3.5-turbo"
        
        if not OPENAI_API_KEY:
            return "❌ OPENAI_API_KEY not set in environment variables"
        
        openai.api_key = OPENAI_API_KEY
        
        response = openai.ChatCompletion.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500
        )
        
        return f"🤖 OpenAI Response:\n{response.choices[0].message.content}"
        
    except Exception as e:
        logger.error(f"OpenAI chat error: {e}")
        return f"❌ OpenAI request failed: {str(e)}"

def wikipedia_search(query: Optional[str] = None) -> str:
    """Search Wikipedia"""
    try:
        if not WIKIPEDIA_AVAILABLE:
            return "❌ Wikipedia not installed. Run: pip install wikipedia"
        
        nlp = extract_params(query, "wikipedia_search") if query else {}
        
        search_query = nlp.get("query", "none") if nlp.get("query") != "none" else input("Search Wikipedia for: ")
        sentences = int(nlp.get("sentences", "none")) if nlp.get("sentences") != "none" else 3
        
        result = wikipedia.summary(search_query, sentences=sentences)
        return f"📚 Wikipedia Result for '{search_query}':\n{result}"
        
    except wikipedia.exceptions.DisambiguationError as e:
        return f"❌ Multiple matches found. Be more specific.\nOptions: {', '.join(e.options[:5])}"
    except wikipedia.exceptions.PageError:
        return f"❌ No page found for '{search_query}'"
    except Exception as e:
        logger.error(f"Wikipedia search error: {e}")
        return f"❌ Wikipedia search failed: {str(e)}"

def send_email(query: Optional[str] = None) -> str:
    """Send single email"""
    try:
        nlp = extract_params(query, "send_email") if query else {}
        
        message = nlp.get("message", "none") if nlp.get("message") != "none" else input("Enter the message to send: ")
        recipient_email = nlp.get("recipient_email", "none") if nlp.get("recipient_email") != "none" else input("Enter the recipient's email: ")
        
        # Simple SMTP implementation
        import smtplib
        from email.mime.text import MIMEText
        
        smtp_server = input("SMTP server (e.g., smtp.gmail.com): ") or "smtp.gmail.com"
        smtp_port = int(input("SMTP port (e.g., 587): ") or 587)
        sender_email = input("Your email: ")
        sender_password = input("Your email password/app password: ")
        
        msg = MIMEText(message)
        msg['Subject'] = 'AI Notification'
        msg['From'] = sender_email
        msg['To'] = recipient_email
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        
        return f"✅ Email sent to {recipient_email}"
        
    except Exception as e:
        logger.error(f"Send email error: {e}")
        return f"❌ Failed to send email: {str(e)}"

def post_tweet(query: Optional[str] = None) -> str:
    """Post a tweet to Twitter/X"""
    try:
        if not TWITTER_AVAILABLE:
            return "❌ Tweepy not installed. Run: pip install tweepy"
        
        if not all([TWITTER_API_KEY, TWITTER_API_SECRET, TWITTER_ACCESS_TOKEN, TWITTER_ACCESS_SECRET]):
            return "❌ Twitter credentials not set in environment variables"
        
        nlp = extract_params(query, "post_tweet") if query else {}
        
        tweet_text = nlp.get("tweet_text", "none") if nlp.get("tweet_text") != "none" else input("Enter tweet text: ")
        
        auth = tweepy.OAuth1UserHandler(
            TWITTER_API_KEY, TWITTER_API_SECRET,
            TWITTER_ACCESS_TOKEN, TWITTER_ACCESS_SECRET
        )
        api = tweepy.API(auth)
        
        api.update_status(status=tweet_text)
        
        return "✅ Tweet posted successfully"
        
    except Exception as e:
        logger.error(f"Post tweet error: {e}")
        return f"❌ Failed to post tweet: {str(e)}"

def schedule_google_calendar_event(self, query: Optional[str] = None) -> str:
    """Schedule Google Calendar event"""
    try:
        if not GOOGLE_API_AVAILABLE:
            return "❌ Google API client not installed. Run: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib"
        
        nlp = self.extract_params(query, "schedule_google_calendar_event") if query else {}
        
        title = nlp.get("title", "none") if nlp.get("title") != "none" else input("Enter the event title: ")
        start_time = nlp.get("start_time", "none") if nlp.get("start_time") != "none" else input("Enter start time (YYYY-MM-DD HH:MM): ")
        end_time = nlp.get("end_time", "none") if nlp.get("end_time") != "none" else input("Enter end time (YYYY-MM-DD HH:MM): ")
        
        if not os.path.exists("google_token.json"):
            return "❌ google_token.json not found. Please authenticate with Google first."
        
        creds = Credentials.from_authorized_user_file(
            "google_token.json", 
            ["https://www.googleapis.com/auth/calendar"]
        )
        service = build("calendar", "v3", credentials=creds)
        
        event = {
            "summary": title,
            "start": {"dateTime": f"{start_time}:00", "timeZone": "UTC"},
            "end": {"dateTime": f"{end_time}:00", "timeZone": "UTC"}
        }
        
        service.events().insert(calendarId="primary", body=event).execute()
        return f"✅ Calendar event '{title}' created successfully"
        
    except Exception as e:
        logger.error(f"Calendar event error: {e}")
        return f"❌ Failed to create calendar event: {str(e)}"

def send_whatsapp_message(self, query: Optional[str] = None) -> str:
    """Send WhatsApp message"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        if not WHATSAPP_TOKEN or not WHATSAPP_PHONE_ID:
            return "❌ WhatsApp credentials not set. Set WHATSAPP_TOKEN and WHATSAPP_PHONE_ID environment variables."
        
        nlp = self.extract_params(query, "send_whatsapp_message") if query else {}
        
        phone = nlp.get("phone", "none") if nlp.get("phone") != "none" else input("Enter phone number (with country code): ")
        message = nlp.get("message", "none") if nlp.get("message") != "none" else input("Enter message: ")
        
        url = f"https://graph.facebook.com/v18.0/{WHATSAPP_PHONE_ID}/messages"
        headers = {
            "Authorization": f"Bearer {WHATSAPP_TOKEN}",
            "Content-Type": "application/json"
        }
        payload = {
            "messaging_product": "whatsapp",
            "to": phone,
            "type": "text",
            "text": {"body": message}
        }
        
        response = requests.post(url, headers=headers, json=payload)
        
        if response.status_code == 200:
            return f"✅ WhatsApp message sent to {phone}"
        else:
            return f"❌ Failed to send message: {response.text}"
            
    except Exception as e:
        logger.error(f"WhatsApp message error: {e}")
        return f"❌ Failed to send WhatsApp message: {str(e)}"

def send_multiple_emails(self, query: Optional[str] = None) -> str:
    """Send emails to multiple recipients"""
    try:
        nlp = self.extract_params(query, "send_multiple_emails") if query else {}
        
        message = nlp.get("message", "none") if nlp.get("message") != "none" else input("Enter message: ")
        emails_string = nlp.get("emails", "none") if nlp.get("emails") != "none" else input("Enter comma-separated emails: ")
        
        emails = [e.strip() for e in emails_string.split(",") if e.strip()]
        
        if not emails:
            return "❌ No valid email addresses provided"
        
        # Simple SMTP implementation
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        smtp_server = input("SMTP server (e.g., smtp.gmail.com): ") or "smtp.gmail.com"
        smtp_port = int(input("SMTP port (e.g., 587): ") or 587)
        sender_email = input("Your email: ")
        sender_password = input("Your email password/app password: ")
        
        results = []
        for email_addr in emails:
            try:
                msg = MIMEMultipart()
                msg['Subject'] = 'AI Notification'
                msg['From'] = sender_email
                msg['To'] = email_addr
                msg.attach(MIMEText(message, 'plain'))
                
                with smtplib.SMTP(smtp_server, smtp_port) as server:
                    server.starttls()
                    server.login(sender_email, sender_password)
                    server.send_message(msg)
                
                results.append(f"✅ Sent to {email_addr}")
            except Exception as e:
                results.append(f"❌ Failed for {email_addr}: {str(e)}")
        
        return "\n".join(results)
        
    except Exception as e:
        logger.error(f"Multiple emails error: {e}")
        return f"❌ Failed to send emails: {str(e)}"

def download_email_attachments(self, query: Optional[str] = None) -> str:
    """Download email attachments"""
    try:
        if not IMAPLIB_AVAILABLE:
            return "❌ IMAP library not available"
        
        nlp = self.extract_params(query, "download_email_attachments") if query else {}
        
        email_user = nlp.get("email_user", "none") if nlp.get("email_user") != "none" else input("Email: ")
        email_pass = nlp.get("email_pass", "none") if nlp.get("email_pass") != "none" else input("Password/App Password: ")
        folder = nlp.get("folder", "none") if nlp.get("folder") != "none" else "attachments"
        
        # Create folder
        os.makedirs(folder, exist_ok=True)
        
        # Connect to Gmail
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(email_user, email_pass)
        mail.select("inbox")
        
        # Search for all emails
        status, data = mail.search(None, "ALL")
        
        email_ids = data[0].split()
        downloaded_count = 0
        
        for email_id in email_ids[:10]:  # Process first 10 emails
            status, msg_data = mail.fetch(email_id, "(RFC822)")
            
            if status != "OK":
                continue
            
            msg = message_from_bytes(msg_data[0][1])
            
            for part in msg.walk():
                if part.get_content_disposition() == 'attachment':
                    filename = part.get_filename()
                    if filename:
                        filepath = os.path.join(folder, filename)
                        with open(filepath, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                        downloaded_count += 1
        
        mail.close()
        mail.logout()
        
        return f"✅ Downloaded {downloaded_count} attachments to '{folder}' folder"
        
    except Exception as e:
        logger.error(f"Email attachments error: {e}")
        return f"❌ Failed to download attachments: {str(e)}"

def send_slack_message(self, query: Optional[str] = None) -> str:
    """Send message to Slack channel"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        if not SLACK_TOKEN:
            return "❌ SLACK_TOKEN not set in environment variables"
        
        nlp = self.extract_params(query, "send_slack_message") if query else {}
        
        message = nlp.get("message", "none") if nlp.get("message") != "none" else input("Enter Slack message: ")
        channel = nlp.get("channel", "none") if nlp.get("channel") != "none" else input("Enter channel (without #): ")
        
        # Slack webhook URL format
        webhook_url = f"https://hooks.slack.com/services/{SLACK_TOKEN}"
        
        payload = {
            "text": message,
            "channel": f"#{channel}",
            "username": "AI Assistant",
            "icon_emoji": ":robot_face:"
        }
        
        response = requests.post(webhook_url, json=payload)
        
        if response.status_code == 200:
            return f"✅ Slack message sent to #{channel}"
        else:
            return f"❌ Failed: {response.text}"
            
    except Exception as e:
        logger.error(f"Slack message error: {e}")
        return f"❌ Failed to send Slack message: {str(e)}"

def github_create_repo(self, query: Optional[str] = None) -> str:
    """Create a GitHub repository"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        if not GITHUB_TOKEN:
            return "❌ GITHUB_TOKEN not set in environment variables"
        
        nlp = self.extract_params(query, "github_create_repo") if query else {}
        
        repo_name = nlp.get("repo_name", "none") if nlp.get("repo_name") != "none" else input("Repository name: ")
        description = nlp.get("description", "none") if nlp.get("description") != "none" else input("Description (optional): ") or ""
        is_private = nlp.get("is_private", "none") == "yes" if nlp.get("is_private") != "none" else input("Private? (yes/no): ").lower() == "yes"
        
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }
        
        data = {
            "name": repo_name,
            "description": description,
            "private": is_private,
            "auto_init": True  # Initialize with README
        }
        
        response = requests.post(
            "https://api.github.com/user/repos",
            headers=headers,
            json=data
        )
        
        if response.status_code == 201:
            repo_url = response.json().get("html_url")
            return f"✅ Repository created: {repo_url}"
        else:
            return f"❌ Failed: {response.json().get('message', 'Unknown error')}"
            
    except Exception as e:
        logger.error(f"GitHub repo creation error: {e}")
        return f"❌ Failed to create repository: {str(e)}"

def azure_vm_control(self, query: Optional[str] = None) -> str:
    """Control Azure Virtual Machines"""
    try:
        # Using Azure CLI via subprocess
        import subprocess
        
        nlp = self.extract_params(query, "azure_vm_control") if query else {}
        
        action = nlp.get("action", "none") if nlp.get("action") != "none" else input("Action (start/stop/restart): ")
        vm_name = nlp.get("vm_name", "none") if nlp.get("vm_name") != "none" else input("VM name: ")
        resource_group = nlp.get("resource_group", "none") if nlp.get("resource_group") != "none" else input("Resource group: ")
        
        if action not in ['start', 'stop', 'restart']:
            return "❌ Invalid action. Use: start, stop, restart"
        
        cmd = f"az vm {action} --name {vm_name} --resource-group {resource_group}"
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        
        if result.returncode == 0:
            return f"✅ Azure VM {vm_name} {action}ed successfully"
        else:
            return f"❌ Failed: {result.stderr}"
            
    except Exception as e:
        logger.error(f"Azure VM control error: {e}")
        return f"❌ Failed to control Azure VM: {str(e)}"

def dropbox_upload(self, query: Optional[str] = None) -> str:
    """Upload file to Dropbox"""
    try:
        if not DROPBOX_AVAILABLE:
            return "❌ Dropbox not installed. Run: pip install dropbox"
        
        if not DROPBOX_TOKEN:
            return "❌ DROPBOX_TOKEN not set in environment variables"
        
        nlp = self.extract_params(query, "dropbox_upload") if query else {}
        
        local_path = nlp.get("local_path", "none") if nlp.get("local_path") != "none" else input("Local file path: ")
        remote_path = nlp.get("remote_path", "none") if nlp.get("remote_path") != "none" else f"/{os.path.basename(local_path)}"
        
        if not os.path.exists(local_path):
            return f"❌ Local file not found: {local_path}"
        
        dbx = dropbox.Dropbox(DROPBOX_TOKEN)
        
        with open(local_path, 'rb') as f:
            dbx.files_upload(f.read(), remote_path, mode=dropbox.files.WriteMode("overwrite"))
        
        return f"✅ File uploaded to Dropbox: {remote_path}"
        
    except Exception as e:
        logger.error(f"Dropbox upload error: {e}")
        return f"❌ Failed to upload to Dropbox: {str(e)}"

def send_telegram_message(self, query: Optional[str] = None) -> str:
    """Send Telegram message"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
            return "❌ TELEGRAM_TOKEN and TELEGRAM_CHAT_ID not set in environment variables"
        
        nlp = self.extract_params(query, "send_telegram_message") if query else {}
        
        message = nlp.get("message", "none") if nlp.get("message") != "none" else input("Enter message: ")
        chat_id = nlp.get("chat_id", "none") if nlp.get("chat_id") != "none" else TELEGRAM_CHAT_ID
        
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "HTML"
        }
        
        response = requests.post(url, json=payload)
        
        if response.status_code == 200:
            return f"✅ Telegram message sent"
        else:
            return f"❌ Failed: {response.text}"
            
    except Exception as e:
        logger.error(f"Telegram message error: {e}")
        return f"❌ Failed to send Telegram message: {str(e)}"

def api_load_test(self, query: Optional[str] = None) -> str:
    """Perform load testing on API endpoints"""
    try:
        if not REQUESTS_AVAILABLE:
            return "❌ Requests module not installed"
        
        nlp = self.extract_params(query, "api_load_test") if query else {}
        
        url = nlp.get("url", "none") if nlp.get("url") != "none" else input("Enter API URL: ")
        users = int(nlp.get("users", "none")) if nlp.get("users") != "none" else 10
        duration = int(nlp.get("duration", "none")) if nlp.get("duration") != "none" else 10
        
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        import threading
        import time
        from collections import defaultdict
        
        stats = defaultdict(int)
        errors = []
        start_time = time.time()
        
        def make_request(thread_id):
            nonlocal stats, errors
            end_time = start_time + duration
            
            while time.time() < end_time:
                try:
                    request_start = time.time()
                    response = requests.get(url, timeout=5)
                    request_time = (time.time() - request_start) * 1000
                    
                    stats['total_requests'] += 1
                    stats['total_time'] += request_time
                    
                    if 200 <= response.status_code < 300:
                        stats['success'] += 1
                    elif 400 <= response.status_code < 500:
                        stats['client_errors'] += 1
                    else:
                        stats['server_errors'] += 1
                    
                    if request_time > 1000:
                        stats['slow_requests'] += 1
                        
                except Exception as e:
                    stats['errors'] += 1
                    errors.append(str(e))
        
        # Create threads
        threads = []
        for i in range(users):
            thread = threading.Thread(target=make_request, args=(i,))
            threads.append(thread)
            thread.start()
        
        # Wait for duration
        time.sleep(duration)
        
        # Wait for threads to finish
        for thread in threads:
            thread.join()
        
        total_time = time.time() - start_time
        
        result = f"🚀 API Load Test Results\n"
        result += "=" * 60 + "\n"
        result += f"🌐 URL: {url}\n"
        result += f"👥 Users: {users}\n"
        result += f"⏱️  Duration: {duration}s\n"
        result += "-" * 60 + "\n"
        result += f"📈 Total Requests: {stats['total_requests']}\n"
        result += f"📊 Requests/Second: {stats['total_requests']/total_time:.2f}\n"
        result += f"✅ Successful: {stats.get('success', 0)}\n"
        result += f"⚠️  Client Errors: {stats.get('client_errors', 0)}\n"
        result += f"❌ Server Errors: {stats.get('server_errors', 0)}\n"
        result += f"💥 Errors: {stats.get('errors', 0)}\n"
        
        if stats['total_requests'] > 0:
            avg_time = stats['total_time'] / stats['total_requests']
            result += f"⚡ Average Response Time: {avg_time:.2f}ms\n"
            success_rate = (stats.get('success', 0) / stats['total_requests']) * 100
            result += f"📊 Success Rate: {success_rate:.1f}%\n"
        
        return result
        
    except Exception as e:
        logger.error(f"API load test error: {e}")
        return f"❌ Load test failed: {str(e)}"